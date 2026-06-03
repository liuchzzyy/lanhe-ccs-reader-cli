"""Validation helpers against the vendor XLSX export."""

from __future__ import annotations

import argparse
import json
from dataclasses import dataclass
from pathlib import Path
from typing import Any

from openpyxl import load_workbook

from .parser import read_ccs


@dataclass(frozen=True)
class ColumnCheck:
    name: str
    xlsx_index: int
    tolerance: float


STRICT_NUMERIC_CHECKS = [
    ColumnCheck("Voltage/V", 8, 0.0),
    ColumnCheck("Current/uA", 9, 1.1e-6),
    ColumnCheck("Capacity/uAh", 10, 7.0e-5),
    ColumnCheck("Energy/uWh", 12, 7.0e-5),
    ColumnCheck("Power/uW", 14, 7.0e-5),
]


DERIVED_NUMERIC_CHECKS = [
    ColumnCheck("dQdV/uAh/V", 15, 5.0e-3),
    ColumnCheck("dVdQ/V/uAh", 16, 5.0e-3),
]


def validate_against_xlsx(
    ccs_path: str | Path,
    xlsx_path: str | Path,
    *,
    timezone: str | None = None,
) -> dict[str, Any]:
    """Compare parsed CCS rows with the official XLSX export."""
    result = read_ccs(ccs_path, timezone=timezone)
    official_rows = _load_official_record_rows(xlsx_path)
    parsed = result.records

    report: dict[str, Any] = {
        "ccs": str(ccs_path),
        "xlsx": str(xlsx_path),
        "record_count": {
            "parsed": int(len(parsed)),
            "xlsx": int(len(official_rows)),
            "match": len(parsed) == len(official_rows),
        },
        "columns": {},
        "derived_columns": {},
        "identity": {},
        "derived_note": (
            "Vendor dQdV and dVdQ are smoothed derived export columns. Raw voltage/current/capacity/energy and "
            "direct power are checked strictly; derivative columns are reported but do not decide ok."
        ),
    }
    if len(parsed) != len(official_rows):
        report["ok"] = False
        return report

    identity_checks = [
        ("Cycle", 0),
        ("Step", 1),
        ("Record", 2),
        ("WorkMode", 3),
        ("StepInProcess", 4),
        ("StepDuration", 5),
        ("StepTime", 6),
        ("TestTime", 7),
        ("SysTime", 19),
        ("Mark1", 20),
        ("Mark2", 21),
    ]
    for name, index in identity_checks:
        parsed_values = parsed[name].astype(str).to_list()
        mismatches = 0
        first_mismatch: dict[str, Any] | None = None
        for row_index, row in enumerate(official_rows):
            parsed_value = parsed_values[row_index]
            official_value = str(row[index])
            if parsed_value != official_value:
                mismatches += 1
                if first_mismatch is None:
                    first_mismatch = {
                        "row": row_index,
                        "record": int(row[2]),
                        "parsed": parsed_value,
                        "xlsx": official_value,
                    }
        report["identity"][name] = {"mismatches": mismatches, "first_mismatch": first_mismatch}

    for check in STRICT_NUMERIC_CHECKS:
        report["columns"][check.name] = _compare_numeric_column(parsed, official_rows, check)

    for check in DERIVED_NUMERIC_CHECKS:
        report["derived_columns"][check.name] = _compare_numeric_column(parsed, official_rows, check)

    report["metadata"] = {
        "test_name": result.metadata.test_name,
        "process_name": result.metadata.process_name,
        "software_version": result.metadata.software_version,
        "serial_number": result.metadata.serial_number,
        "channel_number": result.metadata.channel_number,
        "active_material_mass_g": result.metadata.active_material_mass_g,
        "nominal_specific_capacity_mAh_g": result.metadata.nominal_specific_capacity_mAh_g,
        "data_start_offset": result.metadata.data_start_offset,
        "page_count": result.metadata.page_count,
    }
    report["ok"] = (
        report["record_count"]["match"]
        and all(item["mismatches"] == 0 for item in report["identity"].values())
        and all(item["first_failure"] is None for item in report["columns"].values())
    )
    return report


def _compare_numeric_column(
    parsed: Any,
    official_rows: list[tuple[Any, ...]],
    check: ColumnCheck,
) -> dict[str, Any]:
    """Compare one numeric parsed column with the vendor export rows."""
    parsed_values = [float(value) for value in parsed[check.name].to_list()]
    max_error = 0.0
    mean_error = 0.0
    first_failure: dict[str, Any] | None = None
    for row_index, row in enumerate(official_rows):
        parsed_value = parsed_values[row_index]
        official_value = float(row[check.xlsx_index] or 0.0)
        error = abs(parsed_value - official_value)
        mean_error += error
        if error > max_error:
            max_error = error
        if error > check.tolerance and first_failure is None:
            first_failure = {
                "row": row_index,
                "record": int(row[2]),
                "parsed": parsed_value,
                "xlsx": official_value,
                "error": error,
                "tolerance": check.tolerance,
            }
    return {
        "max_error": max_error,
        "mean_error": mean_error / len(official_rows),
        "tolerance": check.tolerance,
        "first_failure": first_failure,
    }


def build_parser() -> argparse.ArgumentParser:
    parser = argparse.ArgumentParser(description="Validate parsed CCS data against a vendor XLSX export.")
    parser.add_argument("ccs_file", type=Path, help="Path to the .ccs file.")
    parser.add_argument("xlsx_file", type=Path, help="Path to the vendor-exported .xlsx file.")
    parser.add_argument("--timezone", help="IANA timezone for Unix millisecond timestamps, for example Asia/Shanghai.")
    return parser


def main(argv: list[str] | None = None) -> int:
    args = build_parser().parse_args(argv)
    report = validate_against_xlsx(args.ccs_file, args.xlsx_file, timezone=args.timezone)
    print(json.dumps(report, ensure_ascii=False, indent=2))
    return 0 if report.get("ok") else 1


def _load_official_record_rows(xlsx_path: str | Path) -> list[tuple[Any, ...]]:
    wb = load_workbook(xlsx_path, read_only=True, data_only=True)
    rows: list[tuple[Any, ...]] = []
    for ws in wb.worksheets:
        for row in ws.iter_rows(values_only=True):
            if (
                len(row) >= 22
                and isinstance(row[0], int)
                and isinstance(row[1], int)
                and isinstance(row[2], int)
                and isinstance(row[8], (int, float))
            ):
                rows.append(row)
    return rows


if __name__ == "__main__":
    raise SystemExit(main())
